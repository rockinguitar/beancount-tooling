#!/usr/bin/env python3
import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin")
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, ncols: int, color_hex: str) -> None:
    fill = PatternFill(fill_type="solid", start_color=color_hex, end_color=color_hex)
    font = Font(bold=True)
    align = Alignment(horizontal="center")

    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = GRID_BORDER

    ws.freeze_panes = "A2"


def style_rows(ws, start_row: int, end_row: int, ncols: int) -> None:
    zebra_fill = PatternFill(fill_type="solid", start_color="FAFAFA", end_color="FAFAFA")

    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = GRID_BORDER

            # Zebra striping: every second data row
            if r % 2 == 0:
                cell.fill = zebra_fill


def autosize_columns(ws, df: pd.DataFrame) -> None:
    for idx, col_name in enumerate(df.columns, start=1):
        values = [str(col_name)] + df.iloc[:, idx - 1].astype(str).fillna("").tolist()
        ws.column_dimensions[get_column_letter(idx)].width = max(len(v) for v in values) + 2


def convert_one(csv_path: Path, out_dir: Path, sheet_name: str, header_color: str) -> Path:
    df = pd.read_csv(csv_path)

    out_path = out_dir / f"{csv_path.stem}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        ncols = len(df.columns)
        nrows = len(df) + 1  # incl header

        style_header(ws, ncols, header_color)
        style_rows(ws, start_row=2, end_row=nrows, ncols=ncols)
        autosize_columns(ws, df)

        ws.sheet_view.showGridLines = False  # we draw our own borders

    return out_path


def main() -> None:
    p = argparse.ArgumentParser(description="Convert CSV → XLSX with grid & zebra rows.")
    p.add_argument("--in", dest="in_path", default="exports")
    p.add_argument("--out", dest="out_path", default=None)
    p.add_argument("--glob", dest="glob_pat", default="*.csv")
    p.add_argument("--sheet", dest="sheet_name", default="Data")
    p.add_argument("--header-color", dest="header_color", default="E7F3FF")

    args = p.parse_args()

    in_dir = Path(args.in_path)
    out_dir = Path(args.out_path) if args.out_path else in_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    csv_files = sorted(in_dir.glob(args.glob_pat))
    if not csv_files:
        raise SystemExit(f"No files matched: {in_dir}/{args.glob_pat}")

    for csv_file in csv_files:
        out = convert_one(csv_file, out_dir, args.sheet_name, args.header_color)
        print(f"Wrote {out}")


if __name__ == "__main__":
    main()